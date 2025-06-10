from openpyxl import Workbook, workbook, worksheet
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis

from datetime import datetime
import logging

colors = ["FF0000",  # Red
          "00B050",  # Green
          "0070C0",  # Blue
          "7030A0",  # Purple
          "FFC000"]  # Orange

class report:
    def __init__(self):
        pass
        
    def write_to_file(self, path):
        print("base class, don't call this")
        
class xlsx_report(report):
    def __init__(self):
        wb = Workbook()
        data_ws = wb.active
        data_ws.title = "Data"
        self.wb:workbook.workbook = wb
        self.data_ws:worksheet.worksheet.Worksheet = data_ws
        self.x_name = "x-axis"
        self.y_name = "y-axis"
        self.x_range = None
        self.y_range = None
        self.chart_title = "Chart"
        self.column_index = 1
        
    def add_data(self, header, values):
        """
        Adds a column of data with a header.
        :param header: str, name of the column (goes in row 1)
        :param values: list of values (goes in rows 2+)
        """
        col = self.column_index
        self.data_ws.cell(row=1, column=col, value=header)
        for i, val in enumerate(values, start=2):
            if col == 1:
                fmt_val = datetime.strptime(val, "%Y-%m-%d")
            else:
                fmt_val = float(val)
            self.data_ws.cell(row=i, column=col, value=fmt_val)
        self.column_index += 1
        self.row_count = self.data_ws.max_row
    
    def add_chart(self, name:str, columns:list):
        # Create chart
        chart = LineChart()
        chart.title = name

        # X values (no header)
        cats = Reference(
            self.data_ws, 
            min_col=1, 
            min_row=2, 
            max_row=self.row_count)
        
        for column in columns:
            # Y values (with header for title)
            data_ref = Reference(
                self.data_ws, 
                min_col=column, 
                min_row=1, 
                max_row=self.row_count
                )
            # Add data to chart
            chart.add_data(data_ref, titles_from_data=True)
        for i, s in enumerate(chart.series):
            s.graphicalProperties.line.noFill = True  # Remove the line
            s.marker.symbol = "circle"               # Set marker symbol
            s.marker.size = 7                        # Optional: marker size
            color = colors[i % len(colors)]
            s.marker.graphicalProperties.solidFill = color         # Marker color
            s.graphicalProperties.solidFill = color                # Legend marker color
        chart.set_categories(cats)
        chart.x_axis = DateAxis(crossAx=100)
        chart.x_axis.number_format = 'yyyy-mm-dd'
        chart.x_axis.majorTimeUnit = "days"
        chart.width = 32
        chart.height = 12
        # Create a new sheet for the chart
        chart_ws = self.wb.create_sheet(title=name)
        chart_ws.add_chart(chart, "A1")
        
    def save_file(self, path):
        self.wb.save(path)